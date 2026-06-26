#!/usr/bin/env python3
"""
make_history_template.py
========================
One-off generator. Builds CRATEJOY_HISTORY_TEMPLATE.xlsx — the fill-in workbook
we hand to Sheena/Ting so they can supply the past KIT SKU for every box each
of the ~94 missing Cratejoy customers already received.

Pre-fills everything we know from Cratejoy (customer, box #, ship date). The
only blank column they complete is KIT SKU (with a dropdown sourced from our
own kits table).

Reads:   ../CRATEJOY_BACKFILL_LIST.csv   (produced earlier from the live reconcile)
Calls:   Cratejoy API (shipped boxes per customer) + Supabase (kits list) — READ ONLY
Writes:  ../CRATEJOY_HISTORY_TEMPLATE.xlsx

Usage:   python scripts/make_history_template.py
"""
import base64
import csv
import json
import os
import sys
import time
import urllib.request
from pathlib import Path
from urllib.parse import urljoin

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).parent.parent
CSV_IN = ROOT / "CRATEJOY_BACKFILL_LIST.csv"
XLSX_OUT = ROOT / "CRATEJOY_HISTORY_TEMPLATE.xlsx"

load_dotenv()
CID = os.getenv("CRATEJOY_CLIENT_ID", "")
CSEC = os.getenv("CRATEJOY_CLIENT_SECRET", "")
SB_URL = os.getenv("SUPABASE_URL", "")
SB_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_ANON_KEY") or ""
A = base64.b64encode(f"{CID}:{CSEC}".encode()).decode()
CJ = {"Authorization": f"Basic {A}", "Accept": "application/json"}
SB = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}", "Accept": "application/json"}


def get(url, h):
    req = urllib.request.Request(url, headers=h)
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return r.status, json.loads(r.read())
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode("utf-8", "replace")
    except Exception as ex:
        return 0, str(ex)


# 1. read the customer list we already generated
custs = []
with open(CSV_IN, encoding="utf-8") as f:
    for r in csv.DictReader(f):
        custs.append(r)
print(f"customers from list: {len(custs)}")

# 2. per customer, pull SHIPPED boxes (date + cycle), chronological
# NOTE: query by SUBSCRIPTION_ID with pagination — ?customer_id= returns an
# INCOMPLETE set (Cratejoy quirk). subscription_id paginated returns every box
# across all auto-renewed terms (e.g. Rob Nielsen: 18 total, 15 shipped).
box_rows = []
for i, c in enumerate(custs):
    sub = c["subscription_id"]
    cid = c["cratejoy_customer_id"]
    results = []
    url = f"https://api.cratejoy.com/v1/shipments/?subscription_id={sub}&limit=100"
    while url:
        s, d = get(url, CJ)
        if not isinstance(d, dict):
            break
        results.extend(d.get("results", []))
        nx = d.get("next")
        url = urljoin("https://api.cratejoy.com/v1/shipments/", nx) if nx else None
    boxes = []
    for r in results:
        if r.get("status") != "shipped":
            continue
        dt = str(r.get("adjusted_ordered_at") or r.get("shipped_at") or "")[:10]
        cyc = None
        ff = r.get("fulfillments") or []
        if ff:
            cyc = ff[0].get("cycle_number")
        boxes.append((dt, cyc, r.get("id")))
    boxes.sort(key=lambda z: z[0])
    name = f"{c.get('first_name') or ''} {c.get('last_name') or ''}".strip()
    for rank, (dt, cyc, sid) in enumerate(boxes, start=1):
        box_rows.append({
            "name": name,
            "email": c["email"],
            "cj_cust": cid,
            "sub": c["subscription_id"],
            "box_no": rank,
            "ship_date": dt,
            "cj_shipment_id": sid,
        })
    time.sleep(0.07)
    if (i + 1) % 25 == 0:
        print(f"  ...{i + 1}/{len(custs)}  box_rows so far={len(box_rows)}")
print(f"TOTAL box rows to fill: {len(box_rows)}")

# 3. kits reference from Supabase
s, kits = get(
    f"{SB_URL}/rest/v1/kits?select=sku,trimester,is_welcome_kit,size_variant,quantity_available"
    f"&order=trimester,sku&limit=1000",
    SB,
)
kits = kits if isinstance(kits, list) else []
print(f"kits in reference: {len(kits)}")

# 4. build workbook
wb = Workbook()
HDR = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="2F5597")
YHDR = PatternFill("solid", fgColor="FFC000")
YCELL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

# --- Instructions sheet ---
ws = wb.active
ws.title = "Instructions"
ws.column_dimensions["A"].width = 115
lines = [
    ("Oh Baby Boxes - Cratejoy Shipment History (please fill in)", True, 14),
    ("", False, 11),
    ("Why we need this", True, 12),
    ("These Cratejoy subscribers were never synced into the OBB Curation Engine, so the Engine has NO record of which", False, 11),
    ("kits they already received. Before we switch the Engine on for them, we need the kit each box contained - otherwise", False, 11),
    ("the Engine could send a duplicate kit they already have.", False, 11),
    ("", False, 11),
    ("What to do", True, 12),
    ("1. Open the 'Shipment History' tab.", False, 11),
    ("2. There is ONE ROW PER BOX a customer already received. Customer, Email, Box # and Ship Date are pre-filled", False, 11),
    ("   for you (pulled straight from Cratejoy).", False, 11),
    ("3. For each row, fill in ONLY the yellow 'KIT SKU' column = the exact OBB kit that box contained.", False, 11),
    ("   Click the cell and use the dropdown, or type it. Use the 'Available Kits' tab as the reference list.", False, 11),
    ("   Format looks like:  OBB-WK-B3 KITS   or   OBB-CK-41 KITS", False, 11),
    ("4. If you are not sure which kit a box was, leave KIT SKU blank and write a short note in the Notes column.", False, 11),
    ("5. Save the file and send it back to Hasan. He will import it so the Engine knows each customer's full history.", False, 11),
    ("", False, 11),
    ("Notes", True, 12),
    ("- Ship Date is pre-filled from Cratejoy - only change it if you know it is wrong.", False, 11),
    ("- 'Box #' is the order the boxes shipped (1 = the first box that customer ever received).", False, 11),
    ("- Please do NOT delete or reorder rows - the import matches on Email + Ship Date + Kit.", False, 11),
    ("- A customer with 0 boxes shipped will not appear here (nothing to fill in for them).", False, 11),
]
for i, (txt, bold, sz) in enumerate(lines, start=1):
    cc = ws.cell(row=i, column=1, value=txt)
    cc.font = Font(bold=bold, size=sz)
    cc.alignment = WRAP

# --- Available Kits reference sheet ---
wk = wb.create_sheet("Available Kits")
for j, h in enumerate(["KIT SKU", "Trimester", "Welcome Kit?", "Size Variant", "Qty Available"], start=1):
    c = wk.cell(row=1, column=j, value=h)
    c.font = HDR
    c.fill = HFILL
    c.border = BORDER
for r, k in enumerate(kits, start=2):
    wk.cell(row=r, column=1, value=k.get("sku"))
    wk.cell(row=r, column=2, value=f"T{k.get('trimester')}" if k.get("trimester") else "")
    wk.cell(row=r, column=3, value="Yes" if k.get("is_welcome_kit") else "No")
    wk.cell(row=r, column=4, value=k.get("size_variant"))
    wk.cell(row=r, column=5, value=k.get("quantity_available"))
for col, wd in zip("ABCDE", [28, 12, 14, 14, 16]):
    wk.column_dimensions[col].width = wd
wk.freeze_panes = "A2"
kit_last = len(kits) + 1

# --- Shipment History fill-in sheet (placed second, after Instructions) ---
sh = wb.create_sheet("Shipment History", 1)
cols = [
    ("#", 6), ("Customer Name", 24), ("Email", 32), ("Cratejoy Customer ID", 20),
    ("Subscription ID", 18), ("Box #", 8), ("Ship Date (pre-filled)", 18),
    ("KIT SKU  <-- FILL THIS", 26), ("Notes (optional)", 30),
]
for j, (h, wd) in enumerate(cols, start=1):
    c = sh.cell(row=1, column=j, value=h)
    c.border = BORDER
    if h.startswith("KIT SKU"):
        c.fill = YHDR
        c.font = Font(bold=True, color="000000")
    else:
        c.fill = HFILL
        c.font = HDR
    sh.column_dimensions[get_column_letter(j)].width = wd
for i, b in enumerate(box_rows, start=2):
    vals = [i - 1, b["name"], b["email"], b["cj_cust"], b["sub"], b["box_no"], b["ship_date"], None, None]
    for j, v in enumerate(vals, start=1):
        c = sh.cell(row=i, column=j, value=v)
        c.border = BORDER
        if j == 8:
            c.fill = YCELL
sh.freeze_panes = "A2"
if box_rows:
    sh.auto_filter.ref = f"A1:I{len(box_rows) + 1}"
    if kit_last >= 2:
        dv = DataValidation(type="list", formula1=f"='Available Kits'!$A$2:$A${kit_last}", allow_blank=True)
        dv.showErrorMessage = False  # helper dropdown, still allows free text
        sh.add_data_validation(dv)
        dv.add(f"H2:H{len(box_rows) + 1}")

try:
    wb.save(XLSX_OUT)
    saved = XLSX_OUT
except PermissionError:
    saved = XLSX_OUT.with_name("CRATEJOY_HISTORY_TEMPLATE_NEW.xlsx")
    wb.save(saved)
    print("(original file was locked/open in Excel — wrote a new copy instead)")
print(f"\nSaved {saved.name}  ({len(box_rows)} fill-in rows across {len(custs)} customers)")
