#!/usr/bin/env python3
"""
verify_template.py
==================
Full audit: every customer in CRATEJOY_HISTORY_TEMPLATE_NEW.xlsx vs live
Cratejoy API. Checks total shipped box count AND per-subscription cycle
breakdown for customers with multiple subscription terms.

Run: python scripts/verify_template.py
"""
import base64, csv, json, os, sys, time, urllib.request
from collections import defaultdict
from pathlib import Path
from urllib.parse import urljoin

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).parent.parent))
from dotenv import load_dotenv
load_dotenv(dotenv_path=Path(__file__).parent.parent / ".env")

from openpyxl import load_workbook

CID    = os.getenv("CRATEJOY_CLIENT_ID")
CSEC   = os.getenv("CRATEJOY_CLIENT_SECRET")
SB_URL = os.getenv("SUPABASE_URL")
SB_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
cj_h = {"Authorization": "Basic " + base64.b64encode(f"{CID}:{CSEC}".encode()).decode(),
        "Accept": "application/json"}
sb_h = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}", "Accept": "application/json"}


def cj_get(url):
    req = urllib.request.Request(url, headers=cj_h)
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read())
    except Exception:
        return {}


def sb_get(path):
    req = urllib.request.Request(f"{SB_URL}/rest/v1/{path}", headers=sb_h)
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read())


# ── Build email -> cratejoy_customer_id map ───────────────────────────────────
cid_map = {}
with open(Path(__file__).parent.parent / "CRATEJOY_BACKFILL_LIST.csv", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        cid_map[row["email"].strip().lower()] = row["cratejoy_customer_id"].strip()

# Add the 5 old customers (history_pending=true, not in CSV)
old5 = {"sheilajohnson65@gmail.com", "angiepombo69@gmail.com",
        "mhumphrey023@gmail.com", "wauthier.k@gmail.com", "nab8599@gmail.com"}
for c in sb_get("customers?select=email,cratejoy_customer_id&platform=in.(cratejoy,both)&limit=200"):
    e = (c.get("email") or "").lower()
    if e in old5 and c.get("cratejoy_customer_id"):
        cid_map[e] = str(c["cratejoy_customer_id"])

# ── Read template ─────────────────────────────────────────────────────────────
XLSX = Path(__file__).parent.parent / "CRATEJOY_HISTORY_TEMPLATE_NEW.xlsx"
wb = load_workbook(XLSX)
sh = wb["Shipment History"]
template = defaultdict(list)
for row in sh.iter_rows(min_row=2, values_only=True):
    email = (row[2] or "").strip().lower()
    sub   = str(row[4] or "").strip()
    if email:
        template[email].append(sub)

print(f"Template: {len(template)} customers, {sum(len(v) for v in template.values())} rows")
print()

# ── Per-customer verification ─────────────────────────────────────────────────
all_issues = []
count_ok = 0

for i, (email, t_subs) in enumerate(sorted(template.items()), 1):
    cj_cid = cid_map.get(email)
    if not cj_cid:
        all_issues.append(f"NO CJ ID: {email}")
        print(f"[{i:3}] NO_CJ_ID  {email}")
        continue

    t_by_sub = defaultdict(int)
    for s in t_subs:
        t_by_sub[s] += 1

    sd = cj_get(f"https://api.cratejoy.com/v1/subscriptions/?customer_id={cj_cid}&limit=50")
    api_sub_ids = [str(s["id"]) for s in sd.get("results", [])]

    api_by_sub = {}
    for sub_id in api_sub_ids:
        url = f"https://api.cratejoy.com/v1/shipments/?subscription_id={sub_id}&limit=100"
        shipped = []
        seen = set()
        while url:
            d = cj_get(url)
            for r in d.get("results", []):
                if r["id"] not in seen and r.get("status") == "shipped":
                    seen.add(r["id"])
                    ff = r.get("fulfillments") or []
                    cyc = ff[0].get("cycle_number") if ff else "?"
                    tot = ff[0].get("total_cycles") if ff else "?"
                    dt  = str(r.get("adjusted_ordered_at") or r.get("shipped_at") or "")[:10]
                    shipped.append((dt, cyc, tot))
            nx = d.get("next")
            url = urljoin("https://api.cratejoy.com/v1/shipments/", nx) if nx else None
            time.sleep(0.04)
        if shipped:
            api_by_sub[sub_id] = sorted(shipped)

    api_total = sum(len(v) for v in api_by_sub.values())
    t_total   = len(t_subs)
    multi     = len(api_by_sub) > 1

    sub_issues = []
    for sub_id, api_shipped in api_by_sub.items():
        t_count = t_by_sub.get(sub_id, 0)
        if t_count != len(api_shipped):
            sub_issues.append(f"sub {sub_id}: template={t_count} api={len(api_shipped)}")
    for ms in set(api_by_sub.keys()) - set(t_by_sub.keys()):
        sub_issues.append(f"sub {ms}: {len(api_by_sub[ms])} shipped boxes MISSING from template")

    customer_ok = len(sub_issues) == 0 and api_total == t_total

    if customer_ok:
        count_ok += 1
        if multi:
            label = f"MULTI-SUB({len(api_by_sub)} subs)"
            print(f"[{i:3}] OK  {email}  [{label}]")
            for sub_id, shipped in sorted(api_by_sub.items()):
                cycles = ", ".join(f"c{cyc}/{tot}" for _, cyc, tot in shipped)
                print(f"       sub {sub_id}: {len(shipped)} boxes  cycles=[{cycles}]")
    else:
        print(f"[{i:3}] MISMATCH  {email}")
        for si in sub_issues:
            print(f"       {si}")
        all_issues.extend([f"{email}: {si}" for si in sub_issues])
        if api_total != t_total and not sub_issues:
            all_issues.append(f"{email}: total mismatch template={t_total} api={api_total}")

    time.sleep(0.08)

print()
print("=" * 60)
print(f"OK:     {count_ok}/{len(template)}")
print(f"ISSUES: {len(all_issues)}")
if all_issues:
    print()
    for issue in all_issues:
        print(f"  {issue}")
