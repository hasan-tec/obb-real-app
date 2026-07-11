"""
export_kit_sku_list.py — Export all DB kits and their items to Excel
in the OBB Box SKU List format (section label | kit SKUs as headers | items below).

Usage: python scripts/export_kit_sku_list.py
Output: ~/Downloads/OBB_Kit_SKU_List.xlsx
"""

import os
import re
import sys
from pathlib import Path
from collections import defaultdict

from dotenv import load_dotenv

load_dotenv(Path(__file__).parent.parent / ".env")

from supabase import create_client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── Batch code -> month/year label ──────────────────────────────────────────

BATCH_TO_MONTH = {
    # 2021
    "AB": "February 2021",
    # 2022
    "AP": "April 2022",
    # 2023
    "BB": "April 2023",    "BC": "May 2023",      "BD": "June 2023",
    "BI": "November 2023", "BJ": "December 2023",
    # 2024
    "BK": "January 2024",  "BM": "March 2024",    "BN": "April 2024",
    "BO": "May 2024",      "BP": "June 2024",     "BQ": "July 2024",
    "BR": "August 2024",   "BS": "September 2024","BT": "October 2024",
    "BU": "November 2024", "BV": "December 2024",
    # 2025
    "BW": "January 2025",  "BX": "February 2025", "BY": "March 2025",
    "BZ": "April 2025",    "CA": "May 2025",      "CB": "June 2025",
    "CC": "July 2025",     "CD": "August 2025",   "CE": "September 2025",
    "CF": "October 2025",  "CG": "November 2025", "CH": "December 2025",
    # 2026
    "CI": "January 2026",  "CJ": "February 2026", "CK": "March 2026",
    "CL": "April 2026",    "CM": "May 2026",      "CN": "June 2026",
}

# Welcome kit letter -> year group label + sort order
WK_GROUPS = [
    # (letter_chars, section_label, sort_order)
    ({"A", "B"}, "Welcome Kits 2021", 1),
    ({"C", "D"}, "Welcome Kits 2022", 2),
    ({"E", "F"}, "Welcome Kits 2023/24", 3),
    ({"G"},      "Welcome Kits 2025", 4),
    ({"H"},      "Welcome Kits 2026", 5),
]

def wk_label_and_sort(letter: str):
    for letters, label, order in WK_GROUPS:
        if letter.upper() in letters:
            return label, order
    return f"Welcome Kits ({letter or '2020'})", 0


def letter_pos(c: str) -> int:
    return ord(c.upper()) - ord("A") + 1


def batch_age_rank(batch: str) -> int:
    if len(batch) == 2:
        return letter_pos(batch[0]) * 26 + letter_pos(batch[1])
    return 9999


def parse_kit_sku(sku: str):
    """
    Returns (kit_type, batch_key, trimester, size_variant).
    kit_type: 'WK' | 'MONTHLY' | 'OTHER'
    """
    clean = re.sub(r"^OBB-", "", sku)
    clean = re.sub(r"\s+KITS?$", "", clean, flags=re.IGNORECASE).strip()
    clean = clean.replace("-", "").upper()

    # Welcome kit: WK + optional_letter + digit(s)
    m = re.match(r"^WK([A-Z]?)(\d+)$", clean)
    if m:
        letter = m.group(1)
        size = int(m.group(2))
        return ("WK", "WK" + letter, 0, size)

    # Monthly kit: 2-letter code + trimester-digit + size-digit
    m = re.match(r"^([A-Z]{2})(\d)(\d)$", clean)
    if m:
        batch = m.group(1)
        trimester = int(m.group(2))
        size = int(m.group(3))
        return ("MONTHLY", batch, trimester, size)

    return ("OTHER", clean, 0, 0)


def kit_short_name(sku: str) -> str:
    """OBB-CN-41 KITS -> CN41  |  OBB-WK-D2 KITS -> WKD2"""
    clean = re.sub(r"^OBB-", "", sku)
    clean = re.sub(r"\s+KITS?$", "", clean, flags=re.IGNORECASE).strip()
    return clean.replace("-", "").upper()


def fetch_all(db, table: str, select: str):
    """Paginated fetch — safe for tables >1000 rows."""
    rows = []
    offset = 0
    while True:
        batch = db.table(table).select(select).range(offset, offset + 999).execute()
        if not batch.data:
            break
        rows.extend(batch.data)
        if len(batch.data) < 1000:
            break
        offset += 1000
    return rows


def get_display_items(kit_id, kit_items_map, item_by_id, alts_pairs):
    """
    Returns list of display strings for a kit.
    Items that are OR-alternatives of each other (and both present in the kit)
    are merged into one row as 'ITEM_A ORRR ITEM_B'.
    All others sorted alphabetically by name.
    """
    item_ids = set(kit_items_map.get(kit_id, []))
    if not item_ids:
        return []

    used = set()
    or_rows = []

    for (a_id, b_id) in alts_pairs:
        if (a_id in item_ids and b_id in item_ids
                and a_id not in used and b_id not in used):
            or_rows.append((a_id, b_id))
            used.add(a_id)
            used.add(b_id)

    singles = sorted(
        [i for i in item_ids if i not in used],
        key=lambda x: (item_by_id.get(x) or {}).get("name", "")
    )
    or_rows.sort(key=lambda p: (item_by_id.get(p[0]) or {}).get("name", ""))

    rows = []
    for i_id in singles:
        rows.append((item_by_id.get(i_id) or {}).get("name", i_id))
    for (a_id, b_id) in or_rows:
        name_a = (item_by_id.get(a_id) or {}).get("name", a_id)
        name_b = (item_by_id.get(b_id) or {}).get("name", b_id)
        rows.append(f"{name_a} ORRR {name_b}")

    return rows


def main():
    output_path = Path.home() / "Downloads" / "OBB_Kit_SKU_List.xlsx"

    db = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_SERVICE_ROLE_KEY"])

    print("Fetching kits...")
    all_kits = fetch_all(db, "kits", "id, sku, trimester, size_variant, is_welcome_kit, age_rank")
    print(f"  {len(all_kits)} kits")

    print("Fetching items...")
    all_items = fetch_all(db, "items", "id, name, sku")
    item_by_id = {i["id"]: i for i in all_items}
    print(f"  {len(all_items)} items")

    print("Fetching kit_items...")
    all_kit_items = fetch_all(db, "kit_items", "kit_id, item_id")
    print(f"  {len(all_kit_items)} kit_items rows")

    print("Fetching item_alternatives...")
    all_alts = db.table("item_alternatives").select("item_id, alternative_item_id").execute().data or []
    seen = set()
    alts_pairs = []
    for a in all_alts:
        key = tuple(sorted([a["item_id"], a["alternative_item_id"]]))
        if key not in seen:
            seen.add(key)
            alts_pairs.append((a["item_id"], a["alternative_item_id"]))
    print(f"  {len(alts_pairs)} unique alt pairs")

    kit_items_map = defaultdict(list)
    for ki in all_kit_items:
        kit_items_map[ki["kit_id"]].append(ki["item_id"])

    # ── Group kits into sections ───────────────────────────────────────────────

    wk_sections = defaultdict(list)     # label -> [kit, ...]
    wk_section_sort = {}
    monthly_sections = defaultdict(list)  # batch_code -> [kit, ...]
    other_kits = []

    for kit in all_kits:
        kit_type, batch_key, trimester, size = parse_kit_sku(kit["sku"])
        kit["_batch_key"] = batch_key
        kit["_trimester"] = trimester
        kit["_size"] = size

        if kit_type == "WK":
            letter = batch_key[2:]  # "WKA" -> "A"
            label, sort_order = wk_label_and_sort(letter)
            wk_sections[label].append(kit)
            wk_section_sort[label] = sort_order
        elif kit_type == "MONTHLY":
            monthly_sections[batch_key].append(kit)
        else:
            other_kits.append(kit)

    def kit_sort_key(k):
        batch = k.get("_batch_key", "")
        trimester = k.get("_trimester", 0)
        size = k.get("_size", 0)
        # WK kits: group same-letter kits together, then by size variant
        # Monthly kits: by trimester first, then size variant
        if trimester == 0:
            return (0, batch, size)
        return (trimester, "", size)

    for label in wk_sections:
        wk_sections[label].sort(key=kit_sort_key)
    for batch in monthly_sections:
        monthly_sections[batch].sort(key=kit_sort_key)

    # Ordered section lists
    wk_ordered = sorted(wk_sections.keys(), key=lambda l: wk_section_sort.get(l, 99))
    monthly_ordered = sorted(monthly_sections.keys(), key=batch_age_rank)

    # ── Excel styles ──────────────────────────────────────────────────────────

    DARK_FILL  = PatternFill("solid", fgColor="2E4057")   # section header
    TEAL_FILL  = PatternFill("solid", fgColor="048A81")   # kit SKU header
    GRAY_FILL  = PatternFill("solid", fgColor="AAAAAA")   # NO TRI placeholder
    EVEN_FILL  = PatternFill("solid", fgColor="E8F4F8")
    ODD_FILL   = PatternFill("solid", fgColor="FFFFFF")

    WHITE_BOLD = Font(color="FFFFFF", bold=True, size=10)
    GRAY_FONT  = Font(color="FFFFFF", italic=True, size=9)
    ITEM_FONT  = Font(size=10)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kit SKU List"

    current_row = [1]  # mutable so nested func can modify

    def write_section(label, kits_in_section, is_welcome=False):
        r = current_row[0]

        has_t1 = any(k.get("_trimester") == 1 for k in kits_in_section)
        add_placeholder = not has_t1 and not is_welcome

        # Header row: col A = section label, optional col B = NO TRI 1 KIT, then kit SKUs
        header_vals = [label]
        if add_placeholder:
            header_vals.append("NO TRI 1 KIT")
        for k in kits_in_section:
            header_vals.append(kit_short_name(k["sku"]))

        for col_idx, val in enumerate(header_vals, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.alignment = CENTER
            if col_idx == 1:
                cell.fill = DARK_FILL
                cell.font = WHITE_BOLD
            elif add_placeholder and col_idx == 2:
                cell.fill = GRAY_FILL
                cell.font = GRAY_FONT
            else:
                cell.fill = TEAL_FILL
                cell.font = WHITE_BOLD

        ws.row_dimensions[r].height = 24
        r += 1

        # Item rows
        items_lists = [
            get_display_items(k["id"], kit_items_map, item_by_id, alts_pairs)
            for k in kits_in_section
        ]
        max_items = max((len(lst) for lst in items_lists), default=0)
        kit_col_start = 3 if add_placeholder else 2

        for i in range(max_items):
            fill = EVEN_FILL if i % 2 == 0 else ODD_FILL
            ws.cell(row=r, column=1).fill = fill
            if add_placeholder:
                ws.cell(row=r, column=2).fill = fill
            for kit_idx, lst in enumerate(items_lists):
                col = kit_col_start + kit_idx
                val = lst[i] if i < len(lst) else ""
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = fill
                cell.font = ITEM_FONT
                cell.alignment = LEFT
            ws.row_dimensions[r].height = 18
            r += 1

        r += 1  # blank spacer
        current_row[0] = r

    # ── Write sections ────────────────────────────────────────────────────────

    print("\nWriting Excel...")
    for label in wk_ordered:
        write_section(label, wk_sections[label], is_welcome=True)

    for batch in monthly_ordered:
        label = BATCH_TO_MONTH.get(batch, batch)
        write_section(label, monthly_sections[batch], is_welcome=False)

    if other_kits:
        write_section("Other Kits", other_kits, is_welcome=True)

    # ── Column widths ─────────────────────────────────────────────────────────

    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value:
                # Long OR strings cap at 40
                max_len = max(max_len, min(len(str(cell.value)), 40))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 12)

    wb.save(str(output_path))
    print(f"\nSaved: {output_path}")
    print(f"  WK sections   : {len(wk_ordered)}")
    print(f"  Monthly sections: {len(monthly_ordered)}")
    print(f"  Total kits    : {len(all_kits)}")
    print(f"  Total items   : {len(all_items)}")


if __name__ == "__main__":
    main()
