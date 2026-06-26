#!/usr/bin/env python3
"""
fix_old24.py
============
1. Find old pre-existing CJ customers (history_pending=false) that still have
   unshipped boxes in Cratejoy  => flag them history_pending=true
2. Pull their SHIPPED box history => append rows to CRATEJOY_HISTORY_TEMPLATE.xlsx
3. Delete cjtest@example.com from customers table

Run: python scripts/fix_old24.py [--live]  (default: dry-run)
"""
import base64, json, os, sys, time, urllib.request
from pathlib import Path
from urllib.parse import urljoin

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).parent.parent))
from dotenv import load_dotenv
load_dotenv(dotenv_path=Path(__file__).parent.parent / ".env")

LIVE = "--live" in sys.argv
print(f"MODE: {'LIVE' if LIVE else 'DRY-RUN (pass --live to commit changes)'}\n")

SB_URL = os.getenv("SUPABASE_URL")
SB_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
CID    = os.getenv("CRATEJOY_CLIENT_ID")
CSEC   = os.getenv("CRATEJOY_CLIENT_SECRET")

sb_h = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}", "Accept": "application/json",
        "Content-Type": "application/json"}
cj_h = {"Authorization": "Basic " + base64.b64encode(f"{CID}:{CSEC}".encode()).decode(),
        "Accept": "application/json"}


def sb_req(method, path, body=None):
    data = json.dumps(body).encode() if body else None
    req = urllib.request.Request(f"{SB_URL}/rest/v1/{path}", data=data, headers=sb_h, method=method)
    req.add_header("Prefer", "return=representation")
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read())


def cj_get(url):
    req = urllib.request.Request(url, headers=cj_h)
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            return json.loads(r.read())
    except Exception as e:
        return {"error": str(e)}


# ── 1. Fetch old 24 from DB (CJ platform, history_pending=false) ────────────
# Use server-side filters to avoid Python-side ambiguity
old_24 = sb_req("GET",
    "customers?select=id,email,first_name,last_name,subscription_status,"
    "cratejoy_customer_id,history_pending,platform"
    "&platform=in.(cratejoy,both)&history_pending=eq.false&limit=200")
print(f"Old pre-existing customers (history_pending=false): {len(old_24)}")

# ── 2. Check unshipped boxes for each ───────────────────────────────────────
still_shipping = []
dormant = []

for c in old_24:
    email = c["email"]
    cj_cust_id = c.get("cratejoy_customer_id")
    if not cj_cust_id:
        print(f"  NO_CJ_ID   | {email} | {c.get('subscription_status')}")
        dormant.append(c)
        continue
    # Look up their active subscriptions via Cratejoy
    subs_d = cj_get(f"https://api.cratejoy.com/v1/subscriptions/?customer_id={cj_cust_id}&limit=50")
    subs = subs_d.get("results", [])
    if not subs:
        print(f"  NO_SUBS    | {email} | cj_cust={cj_cust_id}")
        dormant.append(c)
        time.sleep(0.1)
        continue
    # Sum unshipped across all their subscriptions
    total_unshipped = 0
    sub_ids = []
    for sub in subs:
        sub_id = sub.get("id")
        sub_ids.append(sub_id)
        d = cj_get(f"https://api.cratejoy.com/v1/shipments/?subscription_id={sub_id}&status=unshipped&limit=50")
        total_unshipped += d.get("count", 0)
        time.sleep(0.05)
    c["_sub_ids"] = sub_ids  # stash for later history fetch
    print(f"  {'SHIPPING' if total_unshipped > 0 else 'DORMANT ':8} | unshipped={total_unshipped:2} | {email} | subs={sub_ids}")
    if total_unshipped > 0:
        still_shipping.append(c)
    else:
        dormant.append(c)
    time.sleep(0.1)

print(f"\nStill-shipping (need history_pending flag): {len(still_shipping)}")
print(f"Dormant (no unshipped boxes):               {len(dormant)}")

# ── 3. Flag still-shipping as history_pending=true ──────────────────────────
print("\n── Flagging still-shipping as history_pending=true ──")
for c in still_shipping:
    print(f"  {'[LIVE] PATCH' if LIVE else '[DRY]  would patch'} {c['email']}")
    if LIVE:
        sb_req("PATCH", f"customers?id=eq.{c['id']}", {"history_pending": True})

# ── 4. Delete cjtest@example.com ────────────────────────────────────────────
print("\n── Deleting test account cjtest@example.com ──")
test_accounts = [c for c in old_24 if c["email"] == "cjtest@example.com"]
for t in test_accounts:
    print(f"  {'[LIVE] DELETE' if LIVE else '[DRY]  would delete'} {t['email']} (id={t['id']})")
    if LIVE:
        req = urllib.request.Request(
            f"{SB_URL}/rest/v1/customers?id=eq.{t['id']}",
            headers=sb_h, method="DELETE")
        with urllib.request.urlopen(req, timeout=15) as r:
            r.read()

# ── 5. Pull shipped history for still-shipping customers ────────────────────
print("\n── Fetching shipped box history for still-shipping customers ──")
box_rows = []
for c in still_shipping:
    cid   = c.get("cratejoy_customer_id", "")
    name  = f"{c.get('first_name','') or ''} {c.get('last_name','') or ''}".strip()
    email = c["email"]
    sub_ids = c.get("_sub_ids", [])

    all_results = []
    for sub_id in sub_ids:
        url = f"https://api.cratejoy.com/v1/shipments/?subscription_id={sub_id}&limit=100"
        while url:
            d = cj_get(url)
            all_results.extend(d.get("results", []))
            nx = d.get("next")
            url = urljoin("https://api.cratejoy.com/v1/shipments/", nx) if nx else None
        time.sleep(0.05)

    shipped = [r for r in all_results if r.get("status") == "shipped"]
    shipped_sorted = sorted(shipped, key=lambda r: str(r.get("adjusted_ordered_at") or r.get("shipped_at") or ""))
    print(f"  {email}: {len(shipped)} shipped boxes across {len(sub_ids)} subs")
    for rank, r in enumerate(shipped_sorted, start=1):
        dt = str(r.get("adjusted_ordered_at") or r.get("shipped_at") or "")[:10]
        # find which sub this shipment belongs to
        ff = r.get("fulfillments") or []
        r_sub = ff[0].get("subscription_id") if ff else (sub_ids[0] if sub_ids else "")
        box_rows.append({
            "name": name,
            "email": email,
            "cj_cust": cid,
            "sub": r_sub,
            "box_no": rank,
            "ship_date": dt,
            "cj_shipment_id": r.get("id"),
        })
    time.sleep(0.1)

print(f"\nTotal shipped box rows to add to template: {len(box_rows)}")

# ── 6. Append rows to existing CRATEJOY_HISTORY_TEMPLATE.xlsx ───────────────
if not box_rows:
    print("No rows to append — done.")
    sys.exit(0)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

THIN  = Side(style="thin", color="BBBBBB")
BORD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
YCELL = PatternFill("solid", fgColor="FFF2CC")

XLSX = Path(__file__).parent.parent / "CRATEJOY_HISTORY_TEMPLATE.xlsx"
if not XLSX.exists():
    XLSX = Path(__file__).parent.parent / "CRATEJOY_HISTORY_TEMPLATE_NEW.xlsx"

wb = load_workbook(XLSX)
sh = wb["Shipment History"]
last_row = sh.max_row
existing_data_rows = last_row - 1  # minus header row

print(f"  Existing data rows in template: {existing_data_rows}")
print(f"  Will append {len(box_rows)} rows starting at row {last_row + 1}")

if LIVE:
    for i, b in enumerate(box_rows, start=1):
        row_num = last_row + i
        row_no  = existing_data_rows + i
        vals = [row_no, b["name"], b["email"], b["cj_cust"], b["sub"],
                b["box_no"], b["ship_date"], None, None]
        for j, v in enumerate(vals, start=1):
            c = sh.cell(row=row_num, column=j, value=v)
            c.border = BORD
            if j == 8:
                c.fill = YCELL

    # Extend data validation to cover new rows
    end_row = last_row + len(box_rows)
    for dv in sh.data_validations.dataValidation:
        if "H" in str(dv.sqref):
            dv.sqref = f"H2:H{end_row}"
            break

    try:
        wb.save(XLSX)
        print(f"  Saved: {XLSX.name}")
    except PermissionError:
        alt = XLSX.with_name("CRATEJOY_HISTORY_TEMPLATE_UPDATED.xlsx")
        wb.save(alt)
        print(f"  (file locked) Saved to: {alt.name}")
else:
    print("[DRY] Would append rows — run with --live to save")

print("\nDone.")
