#!/usr/bin/env python3
"""
import_cratejoy_history.py  (Cratejoy backfill — PHASE 2)
=========================================================
Reads the filled-in Cratejoy shipment-history workbook that Sheena/Ting return
and writes the real past kit history into Supabase, then clears the
`customers.history_pending` flag so the daily Cratejoy sync can start creating
decisions for those subscribers.

WHY THIS EXISTS
---------------
The ~94 backlog Cratejoy subscribers were imported as customers-only with
history_pending=TRUE (see import_cratejoy_customers.py). The daily sync SKIPS
history_pending customers so it can't assign a duplicate welcome kit to someone
who already received boxes off-radar. This script imports the boxes they already
got — which does two things at once:

  1. Gives the engine an accurate received-kit history (so assign_kit won't
     re-send a kit they already have), and
  2. Makes `_compute_order_type()` (app.py) correctly return 'renewal' for them.
     That function's rule is:  renewal <=> >=1 real shipment with ship_date < today.
     So once their past boxes are in `shipments`, every future decision the daily
     sync creates is automatically labelled 'renewal' — never 'new'. There is NO
     separate order_type flag to set here; importing the history IS the fix.

ORDER-TYPE SAFETY (what the dry-run verifies)
---------------------------------------------
For every customer we import, the dry-run prints the order_type the engine WILL
compute after import (renewal if any imported box ship_date < today, else new),
using the exact same rule as _compute_order_type. This lets us confirm — before
touching the DB — that a customer with shipment history is seen as a renewal and
never mislabelled 'new'.

history_pending is cleared ONLY for customers who got >=1 shipment imported.
A customer whose KIT SKU cells are all blank stays history_pending=TRUE and is
listed under NEEDS-FOLLOW-UP so we can chase the missing kit from Sheena.

USAGE
  python scripts/import_cratejoy_history.py --dry-run
  python scripts/import_cratejoy_history.py --dry-run --verbose
  python scripts/import_cratejoy_history.py --live
  python scripts/import_cratejoy_history.py --dry-run --file "C:\\path\\to\\workbook.xlsx"
"""
import argparse
import logging
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client, Client

# ---------------------------------------------------------------------------
# Environment
# ---------------------------------------------------------------------------
load_dotenv()
SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_ANON_KEY") or ""

SCRIPT_DIR = Path(__file__).parent
ROOT = SCRIPT_DIR.parent
# Default to the file Sheena returned in Downloads; override with --file.
DEFAULT_XLSX = Path.home() / "Downloads" / "UPDATED_cratejoy_history_template_new.xlsx"

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(SCRIPT_DIR / "import_cratejoy_history.log", mode="w", encoding="utf-8"),
    ],
)
logger = logging.getLogger("obb_cj_history")


# ---------------------------------------------------------------------------
# Kit cache (mirrors import_wk_history.py so SKUs resolve identically)
# ---------------------------------------------------------------------------
class KitCache:
    def __init__(self, kits: list, kit_items: list):
        self.by_sku: dict = {}
        self.by_normalized: dict = {}
        self.items_by_kit: dict = {}
        for kit in kits:
            db_sku = kit["sku"]
            self.by_sku[db_sku] = {"id": kit["id"], "trimester": kit["trimester"], "sku": db_sku}
            normalized = re.sub(r"\s+KITS?$", "", db_sku, flags=re.IGNORECASE).strip().upper()
            if normalized:
                self.by_normalized[normalized] = db_sku
        for ki in kit_items:
            self.items_by_kit.setdefault(ki["kit_id"], []).append(ki["item_id"])

    def get_kit(self, sku: str) -> Optional[dict]:
        if not sku:
            return None
        r = self.by_sku.get(sku)
        if r:
            return r
        canon = self.by_normalized.get(re.sub(r"\s+KITS?$", "", sku, flags=re.IGNORECASE).strip().upper())
        return self.by_sku.get(canon) if canon else None

    def canonical_sku(self, sku: str) -> str:
        kit = self.get_kit(sku)
        return kit["sku"] if kit else sku

    def get_items(self, kit_id: str) -> list:
        return self.items_by_kit.get(kit_id, [])

    def __len__(self):
        return len(self.by_sku)


def load_kit_cache(db: Client) -> KitCache:
    logger.info("[SETUP] Loading kit cache from Supabase...")
    kits = db.table("kits").select("id, sku, trimester").execute()
    kit_items = db.table("kit_items").select("kit_id, item_id").execute()
    cache = KitCache(kits.data or [], kit_items.data or [])
    logger.info("[SETUP] Kit cache: %d kits loaded", len(cache))
    return cache


def load_all_customers(db: Client) -> dict:
    """{email.lower(): {id, history_pending, cratejoy_customer_id, ...}}"""
    logger.info("[SETUP] Pre-loading existing customers...")
    existing: dict = {}
    offset = 0
    while True:
        rows = (db.table("customers")
                .select("id, email, history_pending, cratejoy_customer_id")
                .range(offset, offset + 999).execute().data or [])
        for r in rows:
            if r.get("email"):
                existing[r["email"].lower()] = r
        if len(rows) < 1000:
            break
        offset += 1000
    logger.info("[SETUP] %d existing customers loaded", len(existing))
    return existing


def load_existing_ship_keys(db: Client) -> set:
    """Dedup set: (customer_id, kit_sku, 'YYYY-MM')."""
    logger.info("[SETUP] Pre-loading existing shipment dedup keys...")
    keys: set = set()
    offset = 0
    while True:
        rows = (db.table("shipments")
                .select("customer_id, kit_sku, ship_date")
                .range(offset, offset + 999).execute().data or [])
        for r in rows:
            if not r.get("ship_date") or not r.get("kit_sku"):
                continue
            keys.add((r["customer_id"], r["kit_sku"], r["ship_date"][:7]))
        if len(rows) < 1000:
            break
        offset += 1000
    logger.info("[SETUP] %d existing shipment keys", len(keys))
    return keys


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------
def parse_ship_date_cell(cell) -> Optional[date]:
    """Workbook ship dates arrive as datetime (new section) or 'YYYY-MM-DD' str."""
    if cell is None:
        return None
    if isinstance(cell, datetime):
        return cell.date()
    if isinstance(cell, date):
        return cell
    s = str(cell).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Last resort: ISO timestamp
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def normalize_kit_sku(raw: str) -> str:
    """Uppercase, collapse spaces, ensure a single ' KITS' suffix if a 'KIT(S)' is present."""
    s = str(raw).strip().upper()
    s = re.sub(r"\s+", " ", s)
    # 'OBB-WK-C1KITS' -> 'OBB-WK-C1 KITS'
    s = re.sub(r"(?<! )KITS?$", lambda m: " " + m.group(0), s)
    return s


def trimester_from_sku(sku: str) -> Optional[int]:
    """Last '-' segment, first digit. OBB-CK-41 KITS -> 4 ; OBB-WK-C3 KITS -> 3."""
    s = re.sub(r"\s+KITS?$", "", sku, flags=re.IGNORECASE).strip()
    parts = s.split("-")
    for ch in (parts[-1] if parts else ""):
        if ch.isdigit():
            t = int(ch)
            return t if 1 <= t <= 4 else None
    return None


# ---------------------------------------------------------------------------
# Workbook reader
# ---------------------------------------------------------------------------
def read_workbook(xlsx_path: Path) -> tuple:
    """
    Returns (rows, blanks) where:
      rows   = [{email, name, cj_cust, sub, box_no, ship_date(date), kit_sku(str)}]
      blanks = [{email, name, box_no, row_num}]  (email present, KIT SKU empty)
    Banner / sub-header / spacer rows (no '@' in email column) are skipped.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    sheet_name = "Shipment History" if "Shipment History" in wb.sheetnames else wb.sheetnames[0]
    sh = wb[sheet_name]
    logger.info("[READ] %s -> sheet '%s' (%s)", xlsx_path.name, sheet_name, sh.dimensions)

    rows, blanks = [], []
    for i, row in enumerate(sh.iter_rows(min_row=2, values_only=True), start=2):
        email = str(row[2] or "").strip().lower() if len(row) > 2 else ""
        if "@" not in email:
            continue  # banner, sub-header, spacer
        name = str(row[1] or "").strip()
        cj_cust = str(row[3] or "").strip()
        sub = str(row[4] or "").strip()
        box_no = row[5] if len(row) > 5 else None
        ship_date = parse_ship_date_cell(row[6] if len(row) > 6 else None)
        kit_raw = str(row[7] or "").strip() if len(row) > 7 else ""

        if not kit_raw:
            blanks.append({"email": email, "name": name, "box_no": box_no, "row_num": i})
            continue

        rows.append({
            "email": email,
            "name": name,
            "cj_cust": cj_cust,
            "sub": sub,
            "box_no": box_no,
            "ship_date": ship_date,
            "kit_sku": normalize_kit_sku(kit_raw),
            "row_num": i,
        })
    logger.info("[READ] %d fill-in shipment rows, %d blank-KIT rows", len(rows), len(blanks))
    return rows, blanks


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Import Cratejoy kit history from Sheena's filled workbook.")
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--dry-run", action="store_true", help="Parse + validate + report. Write NOTHING.")
    g.add_argument("--live", action="store_true", help="Write shipments and clear history_pending.")
    ap.add_argument("--verbose", action="store_true", help="Per-row logging.")
    ap.add_argument("--file", help="Path to the filled workbook (defaults to Downloads).")
    args = ap.parse_args()

    xlsx_path = Path(args.file) if args.file else DEFAULT_XLSX

    logger.info("=" * 68)
    logger.info("  CRATEJOY HISTORY IMPORT (PHASE 2)  —  %s", "DRY RUN" if args.dry_run else "LIVE")
    logger.info("  Workbook: %s", xlsx_path)
    logger.info("=" * 68)

    if not xlsx_path.exists():
        logger.error("Workbook not found: %s", xlsx_path)
        sys.exit(1)
    if not SUPABASE_URL or not SUPABASE_KEY:
        logger.error("SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY not set in .env")
        sys.exit(1)

    db = create_client(SUPABASE_URL, SUPABASE_KEY)

    kit_cache = load_kit_cache(db)
    customers = load_all_customers(db)
    existing_keys = load_existing_ship_keys(db)

    rows, blanks = read_workbook(xlsx_path)

    today = date.today()

    # --- Resolve each row -> customer, dedup, build shipment records ---------
    to_insert: list = []               # shipment records ready for DB
    local_dedup: set = set()
    unmatched_emails: dict = {}         # email -> count of rows
    unknown_kits: set = set()
    no_ship_date: list = []
    dup_skipped = 0
    # Per-customer bookkeeping for history_pending + order_type prediction
    per_customer_import: dict = {}      # customer_id -> {email, imported, past_dated}

    for r in rows:
        email = r["email"]
        cust = customers.get(email)
        if not cust:
            unmatched_emails[email] = unmatched_emails.get(email, 0) + 1
            continue
        cust_id = cust["id"]

        if r["ship_date"] is None:
            no_ship_date.append(r)
            continue

        canonical = kit_cache.canonical_sku(r["kit_sku"])
        kit = kit_cache.get_kit(canonical)
        if not kit:
            unknown_kits.add(canonical)

        ym = r["ship_date"].strftime("%Y-%m")
        key = (cust_id, canonical, ym)
        if key in existing_keys or key in local_dedup:
            dup_skipped += 1
            if args.verbose:
                logger.info("  [DUP] %s %s %s — already present", email, canonical, ym)
            continue
        local_dedup.add(key)

        rec: dict = {
            "customer_id": cust_id,
            "kit_sku": canonical,
            "ship_date": r["ship_date"].isoformat(),
            "platform": "cratejoy",
        }
        if kit:
            rec["kit_id"] = kit["id"]
        t = (kit["trimester"] if kit else None) or trimester_from_sku(canonical)
        if t:
            rec["trimester_at_ship"] = t
        if r["sub"]:
            rec["order_id"] = r["sub"]

        to_insert.append(rec)

        pc = per_customer_import.setdefault(cust_id, {"email": email, "imported": 0, "past_dated": 0})
        pc["imported"] += 1
        if r["ship_date"] < today:
            pc["past_dated"] += 1

        if args.verbose:
            logger.info("  [ROW %d] %s box%s kit=%s date=%s T%s",
                        r["row_num"], email, r["box_no"], canonical, r["ship_date"], rec.get("trimester_at_ship", "?"))

    # Emails that still have >=1 blank KIT SKU box — their received-kit history is
    # INCOMPLETE, so they must stay history_pending=TRUE (else assign_kit could
    # re-send a kit they already got in one of the unfilled boxes).
    emails_with_blanks = {b["email"] for b in blanks}

    # Customers eligible to have history_pending cleared: got >=1 shipment imported
    # this run AND have NO blank boxes anywhere in the workbook.
    clearable_ids = {
        cid for cid, pc in per_customer_import.items()
        if pc["imported"] > 0 and pc["email"] not in emails_with_blanks
    }
    held_ids = {
        cid for cid, pc in per_customer_import.items()
        if pc["email"] in emails_with_blanks
    }

    # Blanks: which matched customers still owe us a KIT SKU
    blank_by_customer: dict = {}
    for b in blanks:
        cust = customers.get(b["email"])
        label = f"{b['name']} <{b['email']}>" + ("" if cust else "  [NOT IN DB]")
        blank_by_customer.setdefault(label, []).append(b["box_no"])

    # -------------------- REPORT --------------------
    logger.info("")
    logger.info("-" * 68)
    logger.info("  RESOLUTION SUMMARY")
    logger.info("-" * 68)
    logger.info("  Fill-in rows read:                  %d", len(rows))
    logger.info("  Shipments to INSERT (new):          %d", len(to_insert))
    logger.info("  Duplicates skipped (already in DB): %d", dup_skipped)
    logger.info("  Customers touched:                  %d", len(per_customer_import))
    logger.info("  Customers -> clear history_pending: %d", len(clearable_ids))
    logger.info("  Held back (blank box, stays flagged): %d", len(held_ids))
    logger.info("  Rows with NO ship date:             %d", len(no_ship_date))
    logger.info("  Unmatched emails (not in DB):       %d", len(unmatched_emails))
    logger.info("  Blank-KIT rows (need Sheena):       %d", len(blanks))
    logger.info("  Kit SKUs not found in kits table:   %d", len(unknown_kits))

    # Order-type prediction — the thing Sheena cares about.
    renewal_ct = sum(1 for pc in per_customer_import.values() if pc["past_dated"] > 0)
    new_ct = len(per_customer_import) - renewal_ct
    logger.info("")
    logger.info("  ORDER TYPE the engine WILL compute after import")
    logger.info("  (renewal <=> >=1 imported box ship_date < %s):", today.isoformat())
    logger.info("     renewal: %d customers", renewal_ct)
    logger.info("     new:     %d customers", new_ct)
    if new_ct:
        for cid, pc in per_customer_import.items():
            if pc["past_dated"] == 0:
                logger.info("       [NEW] %s — all imported boxes dated today/future", pc["email"])

    if unknown_kits:
        logger.info("")
        logger.info("  KIT SKUs not in kits table (shipment stored text-only, no item history):")
        for s in sorted(unknown_kits):
            logger.info("     - %s", s)

    if unmatched_emails:
        logger.info("")
        logger.info("  UNMATCHED EMAILS (in workbook, not in customers table — shipments skipped):")
        for e, n in sorted(unmatched_emails.items()):
            logger.info("     - %s  (%d rows)", e, n)

    if blank_by_customer:
        logger.info("")
        logger.info("  NEEDS-FOLLOW-UP — blank KIT SKU (customer stays history_pending=TRUE):")
        for label, boxes in sorted(blank_by_customer.items()):
            logger.info("     - %s  boxes=%s", label, boxes)

    if no_ship_date:
        logger.info("")
        logger.info("  ROWS WITH NO SHIP DATE (skipped):")
        for r in no_ship_date:
            logger.info("     - row %d  %s  kit=%s", r["row_num"], r["email"], r["kit_sku"])

    # -------------------- DRY RUN EXIT --------------------
    if args.dry_run:
        logger.info("")
        logger.info("DRY RUN COMPLETE — nothing written. Re-run with --live to import.")
        logger.info("Log: %s", SCRIPT_DIR / "import_cratejoy_history.log")
        return

    # -------------------- LIVE WRITE --------------------
    logger.info("")
    logger.info("LIVE IMPORT — writing %d shipments, then clearing %d history_pending flags.",
                len(to_insert), len(clearable_ids))

    shipments_created = 0
    ship_items_linked = 0
    batch_size = 100
    for i in range(0, len(to_insert), batch_size):
        batch = to_insert[i:i + batch_size]
        try:
            result = db.table("shipments").insert(batch).execute()
            inserted = result.data or []
            shipments_created += len(inserted)
            items_to_insert = []
            for shp in inserted:
                kid = shp.get("kit_id")
                if kid:
                    for iid in kit_cache.get_items(kid):
                        items_to_insert.append({"shipment_id": shp["id"], "item_id": iid})
            for j in range(0, len(items_to_insert), 100):
                try:
                    db.table("shipment_items").insert(items_to_insert[j:j + 100]).execute()
                except Exception as e_it:
                    logger.warning("[ITEMS] sub-batch failed: %s", e_it)
            ship_items_linked += len(items_to_insert)
            logger.info("  batch %d: %d shipments, %d items", i // batch_size + 1, len(inserted), len(items_to_insert))
        except Exception as e:
            logger.error("[SHIP] batch %d failed, falling back to per-row: %s", i // batch_size + 1, e)
            for rec in batch:
                try:
                    r2 = db.table("shipments").insert(rec).execute()
                    if r2.data:
                        shipments_created += 1
                        sid = r2.data[0]["id"]
                        kid = r2.data[0].get("kit_id")
                        if kid:
                            for iid in kit_cache.get_items(kid):
                                try:
                                    db.table("shipment_items").insert({"shipment_id": sid, "item_id": iid}).execute()
                                    ship_items_linked += 1
                                except Exception:
                                    pass
                except Exception as e2:
                    logger.error("[SHIP] row failed (%s / %s): %s", rec.get("customer_id"), rec.get("kit_sku"), e2)

    # Clear history_pending for customers who now have real history
    cleared = 0
    for cid in clearable_ids:
        try:
            db.table("customers").update({"history_pending": False}).eq("id", cid).execute()
            cleared += 1
        except Exception as e:
            logger.error("[FLAG] failed to clear history_pending for %s: %s", cid, e)

    logger.info("")
    logger.info("=" * 68)
    logger.info("  IMPORT COMPLETE")
    logger.info("  Shipments created:        %d", shipments_created)
    logger.info("  Shipment items linked:    %d", ship_items_linked)
    logger.info("  history_pending cleared:  %d customers", cleared)
    logger.info("  Still history_pending (blank KIT / unmatched): chase Sheena")
    logger.info("=" * 68)
    logger.info("  Next: the daily Cratejoy sync will now create decisions for these")
    logger.info("  customers, labelled 'renewal' via _compute_order_type. Verify with")
    logger.info("  POST /api/cratejoy/daily-sync?dry_run=true")
    logger.info("  Log: %s", SCRIPT_DIR / "import_cratejoy_history.log")


if __name__ == "__main__":
    main()
