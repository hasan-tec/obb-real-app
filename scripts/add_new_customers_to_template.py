#!/usr/bin/env python3
"""
add_new_customers_to_template.py
=================================
Adds the 4 brand-new Cratejoy customers (0 shipped boxes, first box already
processed manually by Sheena) into CRATEJOY_HISTORY_TEMPLATE_NEW.xlsx under a
clearly labelled separate section at the bottom.

Sheena fills in: Ship Date + KIT SKU for the box she already sent manually.
Once she does, Phase-2 import clears history_pending and the engine takes over.

Run: python scripts/add_new_customers_to_template.py
"""
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "CRATEJOY_HISTORY_TEMPLATE_NEW.xlsx"

# The 4 new customers — first box was processed manually by Sheena
NEW_CUSTOMERS = [
    {"name": "Brandon Thomas",   "email": "brandon.thomas9839@gmail.com", "cj_cust": "6796815820", "sub": "6796816168"},
    {"name": "Rhonda Boatner",   "email": "rkboat822@gmail.com",          "cj_cust": "7048625173", "sub": "7048627771"},
    {"name": "Garrison Goodman", "email": "ggoodman53@gmail.com",         "cj_cust": "100518855",  "sub": "7050897213"},
    {"name": "Darlene Frye",     "email": "dmfrye13@gmail.com",           "cj_cust": "7051338311", "sub": "7051338423"},
]

THIN = Side(style="thin",   color="BBBBBB")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
YCELL = PatternFill("solid", fgColor="FFF2CC")  # yellow — KIT SKU
OCELL = PatternFill("solid", fgColor="FCE4D6")  # orange — Ship Date
OHDR  = PatternFill("solid", fgColor="C65911")  # dark orange header
OFILL = PatternFill("solid", fgColor="FF6600")  # bright orange section banner

wb = load_workbook(XLSX)
sh = wb["Shipment History"]
last_row     = sh.max_row
existing_data = last_row - 1  # rows excluding header

# ── Blank spacer ──────────────────────────────────────────────────────────────
spacer = last_row + 2

# ── Section banner (merged, explains what Sheena needs to do) ────────────────
banner_row = spacer + 1
banner_text = (
    "NEW CUSTOMERS — No past shipment history. "
    "Their first box was processed manually. "
    "Please fill in SHIP DATE and KIT SKU for the box you already sent. "
    "The engine will automatically handle them from their second box onwards once this is filled in."
)
cell = sh.cell(row=banner_row, column=1, value=banner_text)
cell.font      = Font(bold=True, color="FFFFFF", size=11)
cell.fill      = OFILL
cell.alignment = Alignment(wrap_text=True, vertical="center")
sh.merge_cells(start_row=banner_row, start_column=1, end_row=banner_row, end_column=9)
sh.row_dimensions[banner_row].height = 52

# ── Sub-header ────────────────────────────────────────────────────────────────
hdr_row = banner_row + 1
hdrs = ["#", "Customer Name", "Email", "Cratejoy Customer ID",
        "Subscription ID", "Box #", "Ship Date  <-- FILL THIS",
        "KIT SKU  <-- FILL THIS", "Notes (optional)"]
for j, txt in enumerate(hdrs, start=1):
    c = sh.cell(row=hdr_row, column=j, value=txt)
    c.font      = Font(bold=True, color="FFFFFF")
    c.fill      = OHDR
    c.border    = BORD
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── One row per new customer ──────────────────────────────────────────────────
for i, cust in enumerate(NEW_CUSTOMERS, start=1):
    row_num = hdr_row + i
    vals = [existing_data + i, cust["name"], cust["email"],
            cust["cj_cust"], cust["sub"], 1, None, None, None]
    for j, v in enumerate(vals, start=1):
        c = sh.cell(row=row_num, column=j, value=v)
        c.border = BORD
        if j == 7:
            c.fill = OCELL
        elif j == 8:
            c.fill = YCELL

# Extend KIT SKU dropdown to cover new rows
end_row = hdr_row + len(NEW_CUSTOMERS)
for dv in sh.data_validations.dataValidation:
    if "H" in str(dv.sqref):
        dv.sqref = f"H2:H{end_row}"
        break

try:
    wb.save(XLSX)
    print(f"Saved: {XLSX.name}")
    print(f"  Banner at row {banner_row}, data rows {hdr_row+1} to {end_row}")
except PermissionError:
    alt = XLSX.with_name("CRATEJOY_HISTORY_TEMPLATE_NEW_v2.xlsx")
    wb.save(alt)
    print(f"(file locked) Saved to: {alt.name}")
